import pandas as pd

from openpyxl import load_workbook

import copy

from typing import List

from typing import Dict

def write_to_file(Jobdata: List[Dict[str, str]]) -> None:
    """Writes all results in xlsx file"""

    with open ('table.xlsx', 'a+') as my_file:
        df = pd.DataFrame(Jobdata)
        df.to_excel('table.xlsx')
        my_file.close()

    """Makes the table look more comfortable"""
    wb = load_workbook('table.xlsx')
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:      
            alignment = copy.copy(cell.alignment)
            alignment.wrapText=True
            alignment.vertical='top'
            cell.alignment = alignment
    wb.save('table.xlsx')